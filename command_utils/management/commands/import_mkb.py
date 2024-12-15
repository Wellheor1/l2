from concurrent.futures import ThreadPoolExecutor

from django.core.management import BaseCommand
from directions.models import Diagnoses
import requests
from django.db import transaction
from openpyxl import load_workbook


def fetch(url):
    page = requests.get(url)
    return page.json()['list']


class Command(BaseCommand):
    help = "Импорт справочника МКБ"

    def add_arguments(self, parser):
        parser.add_argument('path', type=str)
        parser.add_argument('mode', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        mode = kwargs.get("mode")
        code, nsi_id, title = "", "", ""
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Уникальный идентификатор" in cells:
                    title = cells.index("Наименование")
                    code = cells.index("Код МКБ-10")
                    nsi_id = cells.index("Уникальный идентификатор")
                    starts = True
            else:
                r = Diagnoses.objects.filter(code=cells[code], d_type=mode)
                if not r.exists():
                    Diagnoses(
                        d_type=mode,
                        code=cells[code],
                        title=cells[title],
                        nsi_id=cells[nsi_id],
                        hide=False,
                        m_type=2
                    ).save()
                    print('сохранено', cells[code])  # noqa: T001
