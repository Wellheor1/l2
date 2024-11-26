import datetime
from django.core.management.base import BaseCommand
from django.db.models import Q
from openpyxl import load_workbook
import clients.models as clients
from api.directions.views import eds_documents
from directory.models import ParaclinicInputGroups, ParaclinicInputField
from django.db import transaction
from django.utils import timezone
import logging

from users.models import DoctorProfile
import directions.models as directions
from utils.dates import normalize_dots_date
from sys import stdout
from django.http import HttpRequest
import simplejson as json

logger = logging.getLogger("IF")


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл с пациентами
        """
        parser.add_argument('path', type=str)
        parser.add_argument('research_id', type=str)
        parser.add_argument('doctor_profile_id', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        research_id = kwargs["research_id"]
        doctor_profile_id = kwargs["doctor_profile_id"]
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        base_l2 = None
        snils, lastname, name, patronymic, sex, polis, born_date = None, None, None, None, None, None, None
        title_fields = []

        doc_profile = DoctorProfile.objects.filter(pk=doctor_profile_id).first()
        financing_source = directions.IstochnikiFinansirovaniya.objects.filter(title="ОМС", base__internal_type=True).first()

        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "СНИЛС пациента" in cells and "Полис ОМС" in cells:
                    starts = True
                    lastname = cells.index("Фамилия пациента")
                    name = cells.index("Имя пациента")
                    patronymic = cells.index("Отчество пациента")
                    sex = cells.index("Пол пациента")
                    snils = cells.index("СНИЛС пациента")
                    polis = cells.index("Полис ОМС")
                    born_date = cells.index("Дата рождения пациента")
                    base_l2 = clients.CardBase.objects.filter(internal_type=True)[0]
                    title_fields = cells.copy()
            else:
                ind = clients.Document.objects.filter(Q(document_type__title__iexact="СНИЛС", number=cells[snils]) | Q(document_type__title__iexact="Полис ОМС", number=cells[polis])).first()

                if ind:
                    i = ind.individual
                    if clients.Card.objects.filter(individual=i, base=base_l2).exists():
                        card = clients.Card.objects.filter(individual=i, base=base_l2).first()
                    else:
                        card = clients.Card.objects.create(
                            number=clients.Card.next_l2_n(),
                            base=base_l2,
                            individual=i,
                        )
                        stdout.write(f'Добавлена карта: {card}')
                else:
                    ind = clients.Individual.objects.create(
                        family=cells[lastname],
                        name=cells[name],
                        patronymic=cells[patronymic],
                        birthday=datetime.datetime.strptime(cells[born_date], "%Y-%m-%d %H:%M:%S").date(),
                        sex=cells[sex][0],
                    )

                    if cells[snils]:
                        snils_object = clients.DocumentType.objects.get(title__iexact='СНИЛС')
                        clients.Document.objects.create(document_type=snils_object, number=cells[snils], individual=ind)

                    polis_object = clients.DocumentType.objects.get(title__iexact='Полис ОМС')
                    clients.Document.objects.create(document_type=polis_object, number=cells[polis], individual=ind) if cells[polis] else None

                    card = clients.Card.objects.create(
                        individual=ind,
                        number=clients.Card.next_l2_n(),
                        base=base_l2,
                    )
                    stdout.write(f'Добавлена карта: {card}')
                try:
                    with transaction.atomic():
                        direction = directions.Napravleniya.objects.create(
                            client=card,
                            istochnik_f=financing_source,
                            hospital=doc_profile.hospital,
                            total_confirmed=True,
                            last_confirmed_at=timezone.now(),
                            eds_required_signature_types=['Врач', 'Медицинская организация'],
                        )

                        iss = directions.Issledovaniya.objects.create(
                            napravleniye=direction,
                            research_id=research_id,
                            time_confirmation=timezone.now(),
                            time_save=timezone.now(),
                            doc_confirmation=doc_profile,
                            doc_save=doc_profile,
                            doc_confirmation_string=f"{doc_profile.get_fio_parts()}",
                        )
                        for group in ParaclinicInputGroups.objects.filter(research_id=research_id):
                            for f in ParaclinicInputField.objects.filter(group=group):
                                if f.title in title_fields:
                                    current_cells = title_fields.index(f.title)
                                    if f.field_type == 1:
                                        tmp_value = cells[current_cells]
                                        if "." in tmp_value or "-" in tmp_value:
                                            value = tmp_value[:10]
                                        value = normalize_dots_date(value)
                                    else:
                                        value = cells[current_cells]
                                directions.ParaclinicResult(issledovaniye=iss, field=f, field_type=f.field_type, value=value).save()
                        eds_documents_data = json.dumps({"pk": direction.pk})
                        eds_documents_obj = HttpRequest()
                        eds_documents_obj._body = eds_documents_data
                        eds_documents_obj.user = doc_profile.user
                        eds_documents(eds_documents_obj)
                        stdout.write(f'Добавлено направление: {direction.pk}')

                except Exception as e:
                    logger.exception(e)
                    message = "Серверная ошибка"
                    return {"ok": False, "message": message}
