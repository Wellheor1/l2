import datetime
import json

from django.http import HttpRequest
from openpyxl.reader.excel import load_workbook

from api.models import Application
from api.patients.views import patients_search_card
from clients.models import Card, Individual, HarmfulFactor, PatientHarmfullFactor, CardBase
from contracts.models import Company, CompanyDepartment, MedicalExamination
from integration_framework.views import check_enp


def get_background_token():
    application = Application.objects.filter(active=True, is_background_worker=True).first()
    if application:
        bearer_token = f"Bearer {application.key}"
    else:
        new_application = Application(name="background_worker", is_background_worker=True)
        new_application.save()
        bearer_token = f"Bearer {new_application.key}"
    return bearer_token


def search_by_fio(request_obj, family, name, patronymic, birthday):
    patient_card = None
    params_tfoms = {
        "enp": "",
        "family": family,
        "name": name,
        "bd": birthday,
        "check_mode": "l2-enp-full",
    }
    params_internal = {
        "type": CardBase.objects.get(internal_type=True).pk,
        "extendedSearch": True,
        "form": {
            "family": family,
            "name": name,
            "patronymic": patronymic,
            "birthday": birthday,
            "archive": False,
        },
        "limit": 1,
    }
    request_obj._body = params_tfoms
    current_patient = check_enp(request_obj)
    if current_patient.data.get("message"):
        request_obj._body = json.dumps(params_internal)
        data = patients_search_card(request_obj)
        results_json = json.loads(data.content.decode("utf-8"))
        if len(results_json["results"]) > 0:
            patient_card_pk = results_json["results"][0]["pk"]
            patient_card = Card.objects.filter(pk=patient_card_pk).first()
    elif len(current_patient.data["list"]) > 1:
        return patient_card
    else:
        patient_card = Individual.import_from_tfoms(current_patient.data["list"], None, None, None, True)
    return patient_card


def find_and_replace(text, symbol1, symbol2):
    result = []
    for i in range(len(text)):
        if text[i] == symbol1:
            current_text = text[0:i] + symbol2 + text[i + 1 :]
            result.append(current_text)
        elif text[i] == symbol2:
            current_text = text[0:i] + symbol1 + text[i + 1 :]
            result.append(current_text)
    return result


def search_by_possible_fio(request_obj, name, patronymic, birthday, possible_family):
    if not possible_family:
        return None
    patient_card = None
    for i in possible_family:
        current_family = i
        patient_card = search_by_fio(request_obj, current_family, name, patronymic, birthday)
        if patient_card is not None:
            break
    return patient_card


def search_patient(snils_data, request_user, family_data, name_data, patronymic_data, birthday_data):
    patient_card = None
    bearer_token = get_background_token()
    params = {"enp": "", "snils": snils_data, "check_mode": "l2-snils"}
    request_obj = HttpRequest()
    request_obj._body = params
    request_obj.user = request_user
    request_obj.method = "POST"
    request_obj.META["HTTP_AUTHORIZATION"] = bearer_token
    current_patient = None
    if snils_data and snils_data != "None":
        current_patient = check_enp(request_obj)
    if not current_patient or current_patient.data.get("message"):
        patient_card = search_by_fio(request_obj, family_data, name_data, patronymic_data, birthday_data)
        if patient_card is None:
            possible_family = find_and_replace(family_data, "е", "ё")
            patient_card = search_by_possible_fio(request_obj, name_data, patronymic_data, birthday_data, possible_family)
            if patient_card is None:
                return patient_card
    elif current_patient.data.get("patient_data") and type(current_patient.data.get("patient_data")) != list:
        patient_card_pk = current_patient.data["patient_data"]["card"]
        patient_card = Card.objects.filter(pk=patient_card_pk).first()
    else:
        patient_card = Individual.import_from_tfoms(current_patient.data["patient_data"], None, None, None, True)

    return patient_card


def create_patient(family_data, name_data, patronymic_data, birthday_data, gender_data):
    patient_indv = Individual(
        family=family_data,
        name=name_data,
        patronymic=patronymic_data,
        birthday=birthday_data,
        sex=gender_data,
    )
    patient_indv.save()
    patient_card = Card.add_l2_card(individual=patient_indv)
    return patient_card


def find_factors(harmful_factors: list):
    if not harmful_factors:
        return None
    incorrect_factor = []
    harmful_factors_data = []
    for i in harmful_factors:
        current_code = i.replace(" ", "")
        harmful_factor = HarmfulFactor.objects.filter(title=current_code).first()
        if harmful_factor:
            harmful_factors_data.append({"factorId": harmful_factor.pk})
        else:
            incorrect_factor.append(f"{current_code}")

    return harmful_factors_data, incorrect_factor


def add_factors_data(patient_card: Card, position: str, factors_data: list, exam_data: str, company_inn: str, department: str):
    try:
        PatientHarmfullFactor.save_card_harmful_factor(patient_card.pk, factors_data)
        company_obj = Company.objects.filter(inn=company_inn).first()
        department_obj = CompanyDepartment.objects.filter(company_id=company_obj.pk, title=department).first()
        if department_obj:
            patient_card.work_department_db_id = department_obj.pk
        else:
            new_department = CompanyDepartment.save_department(company_obj.pk, department)
            patient_card.work_department_db_id = new_department.pk
        patient_card.work_position = position.strip()
        patient_card.work_place_db = company_obj
        patient_card.save()
        MedicalExamination.save_examination(patient_card, company_obj, exam_data)
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "message": e}


def normalize_med_exam_data(snils: str, fio: str, birthday: str, gender: str, inn_company: str, code_harmful: str, position: str, examination_date: str, department: str):
    result = {
        "snils": None,
        "fio": None,
        "family": None,
        "name": None,
        "patronymic": None,
        "birthday": None,
        "gender": None,
        "inn_company": None,
        "codes_harmful": None,
        "position": None,
        "examination_date": None,
        "department": None,
    }
    if snils and snils != "None":
        result["snils"] = snils.replace("-", "").replace(" ", "")
    if fio and fio != "None":
        fio_data = fio.split(" ")
        fio_data = [value for value in fio_data if value]
        result["fio"] = " ".join(fio_data)
        result["family"] = fio_data[0]
        result["name"] = fio_data[1]
        if len(fio_data) > 2:
            result["patronymic"] = fio_data[2]
    if birthday and birthday != "None":
        result["birthday"] = birthday.split(" ")[0]
    if gender and gender != "None":
        result["gender"] = gender[0].lower()
    if inn_company and inn_company != "None":
        result["inn_company"] = inn_company
    if code_harmful and code_harmful != "None":
        result["codes_harmful"] = code_harmful.split(",")
    if position and position != "None":
        result["position"] = position
    if examination_date and examination_date != "None":
        result["examination_date"] = examination_date.split(" ")[0]
    if department and department != "None":
        result["department"] = department
    return result


def create_error_data(fio, snils, reason):
    fio_local = fio if fio else snils
    if not fio_local:
        fio_local = "Неизвестно"
    result = {"fio": fio_local, "reason": reason}
    return result


def check_date(date):
    if not date:
        return False
    try:
        datetime.datetime.strptime(date, '%Y-%m-%d')
    except ValueError:
        return False
    return True


def validate_med_exam_data(normalize_data: dict, inn_company) -> dict:
    result = {"ok": True, "data": {}}
    fio = normalize_data["fio"]
    snils = normalize_data["snils"]
    fio_local = fio if fio else snils
    errors = []
    if not fio_local:
        result = {"ok": False, "data": {}, "empty": True}
        return result
    if normalize_data["inn_company"] != inn_company:
        errors.append("ИНН организации не совпадает")
    if not check_date(normalize_data["birthday"]):
        errors.append("Дата рождения: неверная/несуществующая дата")
    if not check_date(normalize_data["examination_date"]):
        errors.append("Дата мед. осмотра: неверная/несуществующая дата")
    if not normalize_data["department"]:
        errors.append("Подразделение не указано")
    if not normalize_data["gender"] in ["м", "ж"]:
        errors.append("Пол указан не верно")
    if len(normalize_data["position"]) > 128:
        errors.append("Должность больше 128 символов")

    if errors:
        result = {"ok": False, "data": {"fio": fio_local, "reason": ", ".join(errors)}, "empty": False}

    return result


def check_need_col(cols: list, need_cols: set):
    other_need_cols = set(set(cols) - need_cols)
    if len(other_need_cols) + len(need_cols) != len(cols):
        return False
    return True


def form_01(request_data):
    """
    Загрузка списка на мед. осмотр

    На входе:
    Файл XLSX с ФИО и датами осмотра.
    Структура:
    снилс(snils), фио(fio), дата рождения(birthday), пол(gender), инн организации(inn_company), код вредности(code_harmful)
    должность(position), дата мед. осмотра(examination_date), подразделение(department)

    """
    file = request_data.get("file")
    other_need_data = request_data.get("other_need_data")
    user = request_data.get("user")
    company_inn = other_need_data.get("companyInn")
    columns = [{"field": 'fio', "key": 'fio', "title": 'ФИО', "align": 'left', "width": 250}, {"field": 'reason', "key": 'reason', "title": 'Причина ошибки'}]
    wb = load_workbook(filename=file)
    ws = wb[wb.sheetnames[0]]
    starts = False
    incorrect_patients = []
    snils_idx, fio_idx, birthday_idx, gender_idx, inn_company_idx, code_harmful_idx, position_idx, examination_date_idx, department_idx = (
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    need_col_name = {"снилс", "фио", "дата рождения", "пол", "инн организации", "код вредности", "должность", "дата мед. осмотра", "подразделение"}
    for index, row in enumerate(ws.rows, 1):
        cells = [str(x.value) for x in row]
        if not starts:
            if "код вредности" in cells:
                if not check_need_col(cells, need_col_name):
                    return {"ok": False, "result": {}, "message": "Нет обязательных полей"}
                snils_idx = cells.index("снилс")
                fio_idx = cells.index("фио")
                birthday_idx = cells.index("дата рождения")
                gender_idx = cells.index("пол")
                inn_company_idx = cells.index("инн организации")
                code_harmful_idx = cells.index("код вредности")
                position_idx = cells.index("должность")
                examination_date_idx = cells.index("дата мед. осмотра")
                department_idx = cells.index("подразделение")
                starts = True
        else:
            snils = cells[snils_idx].strip()
            fio = cells[fio_idx].strip()
            birthday = cells[birthday_idx].strip()
            gender = cells[gender_idx].strip()
            inn_company = cells[inn_company_idx].strip()
            code_harmful = cells[code_harmful_idx].strip()
            position = cells[position_idx].strip()
            examination_date = cells[examination_date_idx].strip()
            department = cells[department_idx].strip()

            normalize_row = normalize_med_exam_data(snils, fio, birthday, gender, inn_company, code_harmful, position, examination_date, department)
            check_result = validate_med_exam_data(normalize_row, company_inn)
            if not check_result["ok"] and check_result.get("empty"):
                continue
            if not check_result["ok"]:
                incorrect_patients.append(check_result["data"])
                continue
            if check_result["ok"] and check_result["data"]:
                incorrect_patients.append(check_result["data"])

            patient_card = search_patient(normalize_row["snils"], user, normalize_row["family"], normalize_row["name"], normalize_row["patronymic"], normalize_row["birthday"])
            if patient_card is None:
                patient_card = create_patient(normalize_row["family"], normalize_row["name"], normalize_row["patronymic"], normalize_row["birthday"], normalize_row["gender"])
            harmful_factors_data, incorrect_factor = find_factors(normalize_row["codes_harmful"])
            if incorrect_factor:
                incorrect_patients.append({"fio": normalize_row["fio"], "reason": f"Неверные факторы: {incorrect_factor}"})
            patient_updated = add_factors_data(patient_card, normalize_row["position"], harmful_factors_data, normalize_row["examination_date"], company_inn, normalize_row["department"])
            if not patient_updated["ok"]:
                incorrect_patients.append({"fio": cells[fio_idx], "reason": f"Сохранение не удалось, ошибка: {patient_updated['message']}"})
    result = {
        "colData": columns,
        "data": incorrect_patients,
    }
    return {"ok": True, "result": result, "message": ""}
