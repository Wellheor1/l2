import base64
import datetime
import json

import requests
from django.http import HttpRequest
from openpyxl.reader.excel import load_workbook

from api.models import Application
from api.patients.views import patients_search_card
from clients.models import Card, Individual, HarmfulFactor, PatientHarmfullFactor, CardBase
from contracts.models import PriceName, PriceCoast, Company, CompanyDepartment, MedicalExamination
from directory.models import Researches
from integration_framework.views import check_enp
from laboratory.settings import RMIS_MIDDLE_SERVER_ADDRESS, RMIS_MIDDLE_SERVER_TOKEN


def form_01(request_data):
    """
    Загрузка цен по прайсу

    На входе:
    Файл XLSX с ценами прайса
    Cтруктура:
    Код по прайсу (internal_code Researches), Услуга (title_researches), следующее поле - любое текстовое название прайса (priceCoasts.coast)
    """
    price_id = request_data.get("entity_id")
    file = request_data.get("file")
    price = PriceName.objects.filter(pk=price_id).first()
    if not price:
        return {"ok": False, "result": [], "message": "Такого прайса нет"}
    wb = load_workbook(filename=file)
    ws = wb[wb.sheetnames[0]]
    internal_code_idx, coast_idx = (
        '',
        '',
    )
    starts = False
    for row in ws.rows:
        cells = [str(x.value) for x in row]
        if not starts:
            if "Код по прайсу" in cells:
                internal_code_idx = cells.index("Код по прайсу")
                try:
                    coast_idx = cells.index(price.title)
                except ValueError:
                    return {"ok": False, "result": [], "message": "Название прайса не совпадает"}
                starts = True
        else:
            internal_code = cells[internal_code_idx].strip()
            try:
                coast = float(cells[coast_idx].strip())
            except Exception:
                continue
            if internal_code == "None" or not coast:
                continue
            service = Researches.objects.filter(internal_code=internal_code).first()
            if not service:
                continue
            current_coast = PriceCoast.objects.filter(price_name_id=price.pk, research_id=service.pk).first()
            if current_coast:
                if current_coast.coast != coast:
                    current_coast.coast = coast
                    current_coast.save()
            else:
                new_coast = PriceCoast(price_name_id=price.pk, research_id=service.pk, coast=coast)
                new_coast.save()
    if not starts:
        return {"ok": False, "result": [], "message": "Не найдены колонка 'Код по прайсу' "}
    return {"ok": True, "result": [], "message": ""}


def form_02(request_data):
    """
    Загрузка посещений из файла

    На входе:
    Файл XLSX с посещениями
    Cтруктура:
    номер карты, Заведующий отделением, Отделение, Услуга, Фамилия, Имя, Отчество, Дата рождения, СНИЛС, Диагноз, Дата услуги, Это травма
    """
    file = request_data.get("file")
    wb = load_workbook(filename=file)
    ws = wb[wb.sheetnames[0]]
    card_number_idx, head_department_idx, department_idx, service_idx, family_idx, name_idx, patronymic_idx, birthday_idx, snils_idx, diagnos_idx, service_date_idx, is_travma_idx = (
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
    )
    starts = False
    file_data = []
    for row in ws.rows:
        cells = [str(x.value) for x in row]
        if not starts:
            if "номер карты" in cells:
                card_number_idx = cells.index("номер карты")
                head_department_idx = cells.index("Заведующий отделением")
                department_idx = cells.index("Отделение")
                service_idx = cells.index("Услуга")
                family_idx = cells.index("Фамилия")
                name_idx = cells.index("Имя")
                patronymic_idx = cells.index("Отчество")
                birthday_idx = cells.index("Дата рождения")
                snils_idx = cells.index("СНИЛС")
                diagnos_idx = cells.index("Диагноз")
                service_date_idx = cells.index("Дата услуги")
                is_travma_idx = cells.index("Это травма")
                starts = True
        else:
            tmp_data = {
                "cardNumber": cells[card_number_idx],
                "headDepartment": cells[head_department_idx],
                "department": cells[department_idx],
                "service": cells[service_idx],
                "family": cells[family_idx],
                "name": cells[name_idx],
                "patronymic": cells[patronymic_idx],
                "birthday": cells[birthday_idx],
                "snils": cells[snils_idx],
                "diagnos": cells[diagnos_idx],
                "serviceDate": cells[service_date_idx],
                "isTravma": cells[is_travma_idx],
            }
            file_data.append(tmp_data)
    if not starts:
        return {"ok": False, "result": [], "message": "Не найдена колонка 'номер карты' "}

    json_str = json.dumps(file_data)
    base64_data = base64.b64encode(json_str.encode())
    json_data = {"data": base64_data}
    headers = {"authorization": f"Bearer {RMIS_MIDDLE_SERVER_TOKEN}"}

    response = requests.post(f"{RMIS_MIDDLE_SERVER_ADDRESS}/send-case-visit", json=json_data, headers=headers)
    result = response.json()

    return {"ok": True, "result": [], "message": f"{result}"}


def get_background_token():
    application = Application.objects.filter(active=True, is_background_worker=True).first()
    if application:
        bearer_token = f"Bearer {application.key}"
    else:
        new_application = Application(name="background_worker", is_background_worker=True)
        new_application.save()
        bearer_token = f"Bearer {new_application.key}"
    return bearer_token


def search_by_fio(request_obj, family, name, patronymic, birthday):
    patient_card = None
    params_tfoms = {
        "enp": "",
        "family": family,
        "name": name,
        "bd": birthday,
        "check_mode": "l2-enp-full",
    }
    params_internal = {
        "type": CardBase.objects.get(internal_type=True).pk,
        "extendedSearch": True,
        "form": {
            "family": family,
            "name": name,
            "patronymic": patronymic,
            "birthday": birthday,
            "archive": False,
        },
        "limit": 1,
    }
    request_obj._body = params_tfoms
    current_patient = check_enp(request_obj)
    if current_patient.data.get("message"):
        request_obj._body = json.dumps(params_internal)
        data = patients_search_card(request_obj)
        results_json = json.loads(data.content.decode("utf-8"))
        if len(results_json["results"]) > 0:
            patient_card_pk = results_json["results"][0]["pk"]
            patient_card = Card.objects.filter(pk=patient_card_pk).first()
    elif len(current_patient.data["list"]) > 1:
        return patient_card
    else:
        patient_card = Individual.import_from_tfoms(current_patient.data["list"], None, None, None, True)
    return patient_card


def find_and_replace(text, symbol1, symbol2):
    result = []
    for i in range(len(text)):
        if text[i] == symbol1:
            current_text = text[0:i] + symbol2 + text[i + 1:]
            result.append(current_text)
        elif text[i] == symbol2:
            current_text = text[0:i] + symbol1 + text[i + 1:]
            result.append(current_text)
    return result


def search_by_possible_fio(request_obj, name, patronymic, birthday, possible_family):
    if not possible_family:
        return None
    patient_card = None
    for i in possible_family:
        current_family = i
        patient_card = search_by_fio(request_obj, current_family, name, patronymic, birthday)
        if patient_card is not None:
            break
    return patient_card


def search_patient(snils_data, request_user, family_data, name_data, patronymic_data, birthday_data):
    patient_card = None
    bearer_token = get_background_token()
    params = {"enp": "", "snils": snils_data, "check_mode": "l2-snils"}
    request_obj = HttpRequest()
    request_obj._body = params
    request_obj.user = request_user
    request_obj.method = "POST"
    request_obj.META["HTTP_AUTHORIZATION"] = bearer_token
    current_patient = None
    if snils_data and snils_data != "None":
        current_patient = check_enp(request_obj)
    if not current_patient or current_patient.data.get("message"):
        patient_card = search_by_fio(request_obj, family_data, name_data, patronymic_data, birthday_data)
        if patient_card is None:
            possible_family = find_and_replace(family_data, "е", "ё")
            patient_card = search_by_possible_fio(request_obj, name_data, patronymic_data, birthday_data, possible_family)
            if patient_card is None:
                return patient_card
    elif current_patient.data.get("patient_data") and type(current_patient.data.get("patient_data")) != list:
        patient_card_pk = current_patient.data["patient_data"]["card"]
        patient_card = Card.objects.filter(pk=patient_card_pk).first()
    else:
        patient_card = Individual.import_from_tfoms(current_patient.data["patient_data"], None, None, None, True)

    return patient_card


def create_patient(family_data, name_data, patronymic_data, birthday_data, gender_data):
    patient_indv = Individual(
        family=family_data,
        name=name_data,
        patronymic=patronymic_data,
        birthday=birthday_data,
        sex=gender_data,
    )
    patient_indv.save()
    patient_card = Card.add_l2_card(individual=patient_indv)
    return patient_card


def find_factors(harmful_factors: list):
    if not harmful_factors:
        return None
    incorrect_factor = []
    harmful_factors_data = []
    for i in harmful_factors:
        current_code = i.replace(" ", "")
        harmful_factor = HarmfulFactor.objects.filter(title=current_code).first()
        if harmful_factor:
            harmful_factors_data.append({"factorId": harmful_factor.pk})
        else:
            incorrect_factor.append(f"{current_code}")

    return harmful_factors_data, incorrect_factor


def add_factors_data(patient_card: Card, position: str, factors_data: list, exam_data: str, company_inn: str, department: str):
    try:
        PatientHarmfullFactor.save_card_harmful_factor(patient_card.pk, factors_data)
        company_obj = Company.objects.filter(inn=company_inn).first()
        department_obj = CompanyDepartment.objects.filter(company_id=company_obj.pk, title=department).first()
        if department_obj:
            patient_card.work_department_db_id = department_obj.pk
        else:
            new_department = CompanyDepartment.save_department(company_obj.pk, department)
            patient_card.work_department_db_id = new_department.pk
        patient_card.work_position = position.strip()
        patient_card.work_place_db = company_obj
        patient_card.save()
        MedicalExamination.save_examination(patient_card, company_obj, exam_data)
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "message": e}


def form_03(request_data):
    """
    Загрузка списка на мед. осмотр

    На входе:
    Файл XLSX с ФИО и датами осмотра.
    Структура:
    снилс(snils), фио(fio), дата рождения(birthday), пол(gender), инн организации(inn_company), код вредности(code_harmful)
    должность(position), дата мед. осмотра(examination_date), подразделение(department)

    """
    file = request_data.get("file")
    wb = load_workbook(filename=file)
    ws = wb[wb.sheetnames[0]]
    starts = False
    incorrect_patients = []
    company_inn = request.POST.get("companyInn", None)
    snils, fio, birthday, gender, inn_company, code_harmful, position, examination_date, department = (
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    for index, row in enumerate(ws.rows, 1):
        cells = [str(x.value) for x in row]
        if not starts:
            if "код вредности" in cells:
                snils = cells.index("снилс")
                fio = cells.index("фио")
                birthday = cells.index("дата рождения")
                gender = cells.index("пол")
                inn_company = cells.index("инн организации")
                code_harmful = cells.index("код вредности")
                position = cells.index("должность")
                examination_date = cells.index("дата мед. осмотра")
                department = cells.index("подразделение")
                starts = True
        else:
            if company_inn != cells[inn_company].strip():
                incorrect_patients.append({"fio": cells[fio], "reason": "ИНН организации не совпадает"})
                continue
            snils_data = cells[snils].replace("-", "").replace(" ", "")
            fio_data, family_data, name_data, patronymic_data = None, None, None, None
            if cells[fio] and cells[fio] != "None":
                fio_data = cells[fio].split(" ")
                family_data = fio_data[0]
                name_data = fio_data[1]
                if len(fio_data) > 2:
                    patronymic_data = fio_data[2]
            birthday_data = cells[birthday].split(" ")[0]
            code_harmful_data = cells[code_harmful].split(",")
            exam_data = cells[examination_date].split(" ")[0]
            try:
                datetime.datetime.strptime(birthday_data, '%Y-%m-%d')
                datetime.datetime.strptime(exam_data, '%Y-%m-%d')
            except ValueError as e:
                incorrect_patients.append({"fio": cells[fio], "reason": f"Неверный формат даты/несуществующая дата в файле: {e}"})
                continue
            gender_data = cells[gender][0]
            department_data = cells[department]
            if fio_data is None and snils_data is None:
                incorrect_patients.append({"fio": f"Строка: {index}", "reason": "Отсутствует данные"})
                continue
            patient_card = search_patient(snils_data, request.user, family_data, name_data, patronymic_data, birthday_data)
            if patient_card is None:
                patient_card = create_patient(family_data, name_data, patronymic_data, birthday_data, gender_data)
            harmful_factors_data, incorrect_factor = find_factors(code_harmful_data)
            if incorrect_factor:
                incorrect_patients.append({"fio": cells[fio], "reason": f"Неверные факторы: {incorrect_factor}"})
            patient_updated = add_factors_data(patient_card, cells[position], harmful_factors_data, exam_data, company_inn, department_data)
            if not patient_updated["ok"]:
                incorrect_patients.append({"fio": cells[fio], "reason": f"Сохранение не удалось, ошибка: {patient_updated['message']}"})

    return {"ok": True, "result": incorrect_patients, "message": ""}
