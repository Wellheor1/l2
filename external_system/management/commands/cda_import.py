from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from external_system.models import CdaFields


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл с CDA-titles столбцами:
        NAME
        """
        fp = kwargs["path"]
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False

        title, doc_refferal, extract = '', '', ''
        for row in ws.rows:
            is_doc_refferal = False
            is_extract = False
            cells = [str(x.value) for x in row]
            if not starts:
                if "NAME" in cells:
                    title = cells.index("NAME")
                    doc_refferal = cells.index("doc_refferal")
                    extract = cells.index("extract")
                    starts = True
            else:
                cda = CdaFields.objects.filter(title=cells[title])
                if not cda.exists():
                    if int(doc_refferal) == 1:
                        is_doc_refferal = True
                    if int(extract) == 1:
                        is_extract = True
                    CdaFields(title=cells[title], is_doc_refferal=is_doc_refferal, is_extract=is_extract).save()
                    self.stdout.write(f'сохранено {cells[title]}')
